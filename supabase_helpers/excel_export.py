"""
supabase_helpers/excel_export.py
==================================
Builds the "Cells graphs" workbook matching the client's Google Sheet:

  - OVERALL      : simple aggregate format — one row of TOTAL CELL
                    ATTENDANCE per week, one row of TOTAL NO. OF CELLS per
                    week, plus the embedded dual-line chart. (Matches the
                    original client screenshot exactly — this sheet's
                    format does not change.)
  - MEN          : per-leader breakdown table (leader block: leader name +
                    cell count + weekly active-cells row, each cell's
                    weekly attendance with a TOTAL column, and a weekly
                    totals row per leader). Column A (names) + B (cell
                    count) + C (TOTAL) stay frozen on the left; the sheet
                    scrolls horizontally through the week columns.
  - WOMEN        : same layout as MEN, filtered to female leaders.
  - <Leader>     : one sheet per leader — chart ONLY (no visible table).
                    The leader's weekly totals/active-cells data is still
                    written to the sheet (the chart needs real cell
                    references) but those rows are hidden.

Schema notes (see Stats_ServiceCheckin_Tables.txt / project memory):
  - Table name is `event_sessions` (correct spelling).
  - `event_sessions` join target is `events`.
  - `events.Organization` is capital-O (confirmed via Postgres error hint);
    `events.event_leader` holds the leader's display name
    (e.g. "Ps Gavin Enslin").
  - Supabase Python SDK filters on joined-table columns via dot notation are
    unreliable — all filtering on `events.*` fields is done in Python after
    fetching, never via `.ilike()`/`.eq()` in the query builder.
"""

from __future__ import annotations

import io
import logging
from collections import defaultdict
from datetime import date
from typing import Optional

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Leader rosters — used to split MEN / WOMEN sheets and to know which
# per-leader sheets to generate. Adjust/extend as the org's leader roster
# changes; matching is substring-based against events.event_leader.
# ---------------------------------------------------------------------------
MEN_LEADERS = [
    "Gavin", "Esdras", "Gloire", "John-Luke", "Kenny", "Nash",
    "Ryan", "Shane", "Thabo", "Vusi", "Yannis",
]
WOMEN_LEADERS = [
    "Vicky", "Bernice", "Cynthia", "Denise", "Kayla", "Louange", "Glenda", "Sasha",
]


def _get_openpyxl():
    """Import openpyxl with a clear error if it isn't installed."""
    try:
        from openpyxl import Workbook
        from openpyxl.chart import LineChart, Reference
        from openpyxl.chart.series import SeriesLabel
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        return (
            Workbook, Font, PatternFill, Alignment, Border, Side,
            LineChart, Reference, SeriesLabel, get_column_letter,
        )
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is not installed. Add 'openpyxl>=3.1.0' to requirements.txt "
            "and redeploy."
        ) from e


def _match_leader(full_event_leader_name: str, leader_key: str) -> bool:
    """True if leader_key (e.g. 'Gavin') appears in the leader's stored name
    (e.g. 'Ps Gavin Enslin')."""
    if not full_event_leader_name:
        return False
    return leader_key.strip().lower() in full_event_leader_name.strip().lower()


def _fmt_date(iso: str) -> str:
    """Convert 'YYYY-MM-DD' -> 'DD/MM/YY'."""
    try:
        d = date.fromisoformat(iso[:10])
        return d.strftime("%d/%m/%y")
    except Exception:
        return iso


# ---------------------------------------------------------------------------
# 1. Fetch + filter
# ---------------------------------------------------------------------------

def _get_weekly_cells_data(
    org_filter: Optional[dict],
    start_date: str,
    end_date: str,
) -> list[dict]:
    """
    Fetch every event_sessions row in the date window (joined to events),
    then filter down to Cells-type sessions (and the caller's org, if any)
    in Python — never via SDK filters on the joined `events` columns.
    """
    from supabase_helpers.supabase_connection import supabase as sb

    response = (
        sb.table("event_sessions")
        .select(
            "session_date, week_identifier, checked_in_count, event_id, "
            "is_did_not_meet, "
            "events!inner(event_name, event_type_name, Organization, event_leader)"
        )
        .gte("session_date", start_date)
        .lte("session_date", end_date)
        .execute()
    )
    raw_rows = response.data or []

    logger.info(
        "[excel_export] Raw rows fetched: %d (date range %s -> %s)",
        len(raw_rows), start_date, end_date,
    )
    if raw_rows:
        logger.info("[excel_export] Sample row: %s", raw_rows[0])
    else:
        logger.warning("[excel_export] No rows returned from event_sessions for this date range")

    org_value = (org_filter or {}).get("organization")
    org_value = org_value.strip().lower() if org_value else None

    rows: list[dict] = []
    for row in raw_rows:
        if row.get("is_did_not_meet"):
            continue

        ev = row.get("events") or {}
        event_type = (ev.get("event_type_name") or "").strip().lower()
        if event_type != "cells":
            continue

        if org_value:
            event_org = (ev.get("Organization") or "").strip().lower()
            if event_org != org_value:
                continue

        rows.append({
            "session_date":     row.get("session_date"),
            "week_identifier":  row.get("week_identifier") or "",
            "checked_in_count": row.get("checked_in_count") or 0,
            "event_id":         row.get("event_id"),
            "event_name":       ev.get("event_name") or "Unnamed Cell",
            "event_leader":     ev.get("event_leader") or "Unassigned",
        })

    logger.info("[excel_export] Rows after filtering to Cells: %d", len(rows))
    if raw_rows and not rows:
        logger.warning(
            "[excel_export] All rows dropped by filter - sample event data: %s",
            raw_rows[0].get("events"),
        )

    return rows


def _build_week_columns(rows: list[dict]) -> list[dict]:
    """Sorted list of {week_identifier, label, start} for every week present."""
    weeks: dict[str, list[str]] = defaultdict(list)
    for r in rows:
        if r["week_identifier"] and r["session_date"]:
            weeks[r["week_identifier"]].append(str(r["session_date"])[:10])

    week_cols = []
    for week_id, dates in weeks.items():
        dates_sorted = sorted(dates)
        start, end = dates_sorted[0], dates_sorted[-1]
        label = f"WEEK of {_fmt_date(start)}-{_fmt_date(end)}"
        week_cols.append({"week_identifier": week_id, "label": label, "start": start})

    week_cols.sort(key=lambda w: w["start"])
    return week_cols


def _build_overall_weekly(rows: list[dict]) -> dict[str, dict]:
    """week_identifier -> {"attendance": int, "cells": set(event_id)}"""
    weeks: dict[str, dict] = defaultdict(lambda: {"attendance": 0, "cells": set()})
    for r in rows:
        wk = r["week_identifier"]
        if not wk:
            continue
        weeks[wk]["attendance"] += r["checked_in_count"]
        weeks[wk]["cells"].add(r["event_id"])
    return weeks


def _build_leader_matrix(rows: list[dict]) -> dict:
    """
    leader -> {
        "events": {event_name: {week_identifier: attendance_sum}},
        "week_totals": {week_identifier: total_attendance},
        "week_active_cells": {week_identifier: distinct_cell_count},
    }
    """
    leaders: dict[str, dict] = {}
    leader_week_events: dict[tuple, set] = defaultdict(set)

    for r in rows:
        leader = r["event_leader"]
        event = r["event_name"]
        week = r["week_identifier"]
        count = r["checked_in_count"]

        bucket = leaders.setdefault(leader, {
            "events": defaultdict(lambda: defaultdict(int)),
            "week_totals": defaultdict(int),
            "week_active_cells": defaultdict(int),
        })
        bucket["events"][event][week] += count
        bucket["week_totals"][week] += count
        leader_week_events[(leader, week)].add(event)

    for (leader, week), evset in leader_week_events.items():
        leaders[leader]["week_active_cells"][week] = len(evset)

    return leaders


def _gender_of(name: str) -> Optional[str]:
    if any(_match_leader(name, k) for k in MEN_LEADERS):
        return "men"
    if any(_match_leader(name, k) for k in WOMEN_LEADERS):
        return "women"
    return None


# ---------------------------------------------------------------------------
# 2a. OVERALL sheet — simple aggregate format (unchanged from the original)
# ---------------------------------------------------------------------------

def _write_overall_sheet(
    wb, week_cols, overall_weekly,
    Font, PatternFill, Alignment, Border, Side,
    LineChart, Reference, SeriesLabel, get_column_letter,
):
    ws = wb.create_sheet(title="OVERALL")

    thin       = Side(border_style="thin", color="000000")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    black_fill = PatternFill("solid", fgColor="000000")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    def _hdr(cell, text) -> None:
        cell.value     = text
        cell.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        cell.fill      = black_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    def _data(cell, value) -> None:
        cell.value     = value
        cell.font      = Font(name="Arial", size=10, color="000000")
        cell.fill      = white_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border

    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 14

    _hdr(ws["A1"], "x")
    _hdr(ws["B1"], "No. of cells")
    _hdr(ws["A2"], "TOTAL CELL ATTENDANCE")
    _hdr(ws["A3"], "TOTAL NO. OF CELLS")
    _hdr(ws["B2"], "")
    _hdr(ws["B3"], "")

    for i, week in enumerate(week_cols):
        col_idx    = i + 3
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 14
        wk = week["week_identifier"]
        data = overall_weekly.get(wk, {"attendance": 0, "cells": set()})

        _hdr(ws.cell(row=1, column=col_idx), week["label"])
        _data(ws.cell(row=2, column=col_idx), data["attendance"])
        _data(ws.cell(row=3, column=col_idx), len(data["cells"]))

    if week_cols:
        n              = len(week_cols)
        first_data_col = 3
        last_data_col  = 3 + n - 1

        chart          = LineChart()
        chart.title    = "TOTAL CELL ATTENDANCE"
        chart.width    = 22
        chart.height   = 14
        chart.y_axis.title = "TOTAL CELL ATTENDANCE"
        chart.x_axis.title = "WEEKS CELLS RAN"
        # IMPORTANT: do NOT set chart.style here. Several of openpyxl's
        # built-in numeric chart styles (this includes style 10) render as
        # "markers only, no connecting line" in Excel/Sheets, which is
        # exactly the "scattered dots with no line" bug. We fully control
        # line/marker appearance explicitly per-series below instead.

        att_ref = Reference(ws, min_col=first_data_col, max_col=last_data_col, min_row=2, max_row=2)
        # from_rows=True is critical here: our data is laid out as ONE ROW
        # spanning many week-columns. Without this flag, openpyxl's default
        # (from_rows=False) treats EACH COLUMN as its own separate series —
        # which is exactly why the chart previously showed ~20 one-point
        # "series" (visible as "15 more" in the legend) instead of a single
        # connected line across all weeks.
        chart.add_data(att_ref, titles_from_data=False, from_rows=True)
        s0 = chart.series[0]
        s0.tx = SeriesLabel(v="Total Cell Attendance")
        s0.smooth = False
        s0.marker.symbol = "none"
        s0.graphicalProperties.line.noFill = False
        s0.graphicalProperties.line.solidFill = "4472C4"
        s0.graphicalProperties.line.width     = 20000

        cells_ref = Reference(ws, min_col=first_data_col, max_col=last_data_col, min_row=3, max_row=3)
        chart.add_data(cells_ref, titles_from_data=False, from_rows=True)
        s1 = chart.series[1]
        s1.tx = SeriesLabel(v="No. of Active Cells")
        s1.smooth = False
        s1.marker.symbol = "none"
        s1.graphicalProperties.line.noFill = False
        s1.graphicalProperties.line.solidFill = "FF0000"
        s1.graphicalProperties.line.width     = 20000

        cats = Reference(ws, min_col=first_data_col, max_col=last_data_col, min_row=1)
        chart.set_categories(cats)

        ws.add_chart(chart, "A5")

    return ws


# ---------------------------------------------------------------------------
# 2b. Per-leader breakdown table (MEN / WOMEN / individual leader sheets)
# ---------------------------------------------------------------------------

def _write_header_row(ws, week_cols, header_font, header_fill, header_align, get_column_letter):
    c = ws.cell(row=1, column=1, value="EVENT NAME")
    c.font, c.fill, c.alignment = header_font, header_fill, header_align

    c = ws.cell(row=1, column=2, value="No. of\ncells")
    c.font, c.fill, c.alignment = header_font, header_fill, header_align

    c = ws.cell(row=1, column=3, value="TOTAL")
    c.font, c.fill, c.alignment = header_font, header_fill, header_align

    for i, w in enumerate(week_cols):
        c = ws.cell(row=1, column=4 + i, value=w["label"])
        c.font, c.fill, c.alignment = header_font, header_fill, header_align

    # Freeze row 1 and columns A-C (names / cell-count / total) so the sheet
    # scrolls horizontally through the week columns while the name column
    # stays pinned on the left.
    ws.freeze_panes = "D2"

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 12
    ws.row_dimensions[1].height = 36
    for i in range(len(week_cols)):
        ws.column_dimensions[get_column_letter(4 + i)].width = 16


def _write_leader_block(ws, start_row, leader_name, leader_data, week_cols,
                         label_font, data_align, indent_align):
    """
    Writes one leader's block:
      Row 1 of block: leader name | no. of cells | (blank) | active cells per week
      Rows 2..n:       cell name | (blank) | TOTAL for that cell | weekly attendance
      Final row:       (blank) | (blank) | (blank) | weekly attendance totals

    Returns (next_free_row, active_cells_row, totals_row).
    """
    events = sorted(leader_data["events"].keys())
    row = start_row

    active_cells_row = row
    c = ws.cell(row=row, column=1, value=leader_name)
    c.font = label_font
    c = ws.cell(row=row, column=2, value=len(events))
    c.font, c.alignment = label_font, data_align
    for i, w in enumerate(week_cols):
        c = ws.cell(row=row, column=4 + i,
                     value=leader_data["week_active_cells"].get(w["week_identifier"], 0))
        c.font, c.alignment = label_font, data_align
    row += 1

    for event in events:
        c = ws.cell(row=row, column=1, value=event)
        c.alignment = indent_align

        event_total = sum(leader_data["events"][event].values())
        ws.cell(row=row, column=3, value=event_total).alignment = data_align

        for i, w in enumerate(week_cols):
            val = leader_data["events"][event].get(w["week_identifier"], 0)
            ws.cell(row=row, column=4 + i, value=val).alignment = data_align
        row += 1

    totals_row = row
    for i, w in enumerate(week_cols):
        c = ws.cell(row=row, column=4 + i,
                     value=leader_data["week_totals"].get(w["week_identifier"], 0))
        c.font, c.alignment = label_font, data_align
    row += 2  # blank separator row

    return row, active_cells_row, totals_row


def _add_leader_chart(ws, leader_name, week_cols, anchor_row,
                       active_cells_row, totals_row, LineChart, Reference, SeriesLabel):
    n_weeks = len(week_cols)
    if n_weeks == 0:
        return

    min_col, max_col = 4, 4 + n_weeks - 1

    chart = LineChart()
    chart.title = f"{leader_name} - TOTAL CELL ATTENDANCE"
    chart.x_axis.title = "WEEKS CELLS RAN"
    chart.y_axis.title = "TOTAL CELL ATTENDANCE"
    chart.width, chart.height = 22, 14
    # Same reasoning as the OVERALL chart: no numeric chart.style — we set
    # line/marker appearance explicitly per-series instead.

    # NOTE: rows are no longer marked `hidden` (see _hide_table_rows) — they're
    # shrunk to 1pt height instead, because Google Sheets' .xlsx importer drops
    # chart series whose source cells sit in a hidden row regardless of
    # plotVisOnly. Leaving plotVisOnly at its default (True) is fine now.

    att_ref = Reference(ws, min_col=min_col, max_col=max_col, min_row=totals_row, max_row=totals_row)
    chart.add_data(att_ref, titles_from_data=False, from_rows=True)
    s0 = chart.series[0]
    s0.tx = SeriesLabel(v="Total Cell Attendance")
    s0.smooth = False
    s0.marker.symbol = "none"
    s0.graphicalProperties.line.noFill = False
    s0.graphicalProperties.line.solidFill = "4472C4"
    s0.graphicalProperties.line.width = 20000

    cells_ref = Reference(ws, min_col=min_col, max_col=max_col, min_row=active_cells_row, max_row=active_cells_row)
    chart.add_data(cells_ref, titles_from_data=False, from_rows=True)
    s1 = chart.series[1]
    s1.tx = SeriesLabel(v="No. of Active Cells")
    s1.smooth = False
    s1.marker.symbol = "none"
    s1.graphicalProperties.line.noFill = False
    s1.graphicalProperties.line.solidFill = "FF0000"
    s1.graphicalProperties.line.width = 20000

    cats = Reference(ws, min_col=min_col, max_col=max_col, min_row=1, max_row=1)
    chart.set_categories(cats)

    ws.add_chart(chart, f"A{anchor_row}")


def _hide_table_rows(ws, first_row: int, last_row: int) -> None:
    """
    Collapse the underlying data rows so an individual leader sheet reads as
    chart-only, WITHOUT using the `hidden` row flag.

    Why not `hidden`: openpyxl's `chart.plotVisOnly = False` is an Excel-only
    escape hatch that tells Excel "plot this series even though its source
    row is hidden." Google Sheets' .xlsx importer does not honor
    `plotVisOnly` at all — when a chart's series reference sits inside a row
    marked hidden, Sheets drops the series entirely, which is exactly the
    "Add a series to start visualising your data" empty-chart bug. Since
    these workbooks are opened in Google Sheets, we instead shrink the rows
    to near-zero height (still technically visible, so Sheets keeps the
    chart data) rather than hiding them outright.
    """
    for r in range(first_row, last_row + 1):
        ws.row_dimensions[r].height = 1
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = None


# ---------------------------------------------------------------------------
# 3. Entry point
# ---------------------------------------------------------------------------

def build_cells_excel(
    org_filter: Optional[dict],
    start_date: str,
    end_date: str,
) -> bytes:
    """
    Build and return raw .xlsx bytes:
      - OVERALL: simple aggregate format + chart (unchanged)
      - MEN / WOMEN: per-leader breakdown tables
      - one sheet per named leader: chart only, no visible table
    """
    (
        Workbook, Font, PatternFill, Alignment, Border, Side,
        LineChart, Reference, SeriesLabel, get_column_letter,
    ) = _get_openpyxl()

    rows = _get_weekly_cells_data(org_filter, start_date, end_date)
    week_cols = _build_week_columns(rows)
    overall_weekly = _build_overall_weekly(rows)
    leaders = _build_leader_matrix(rows)
    all_leader_names = sorted(leaders.keys())

    header_font = Font(color="FFFFFF", bold=True, name="Arial")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = Alignment(horizontal="center", vertical="center")
    indent_align = Alignment(indent=1)
    label_font = Font(bold=True)

    wb = Workbook()
    wb.remove(wb.active)

    # -- OVERALL (unchanged simple aggregate format) -------------------------
    _write_overall_sheet(
        wb, week_cols, overall_weekly,
        Font, PatternFill, Alignment, Border, Side,
        LineChart, Reference, SeriesLabel, get_column_letter,
    )

    def new_breakdown_sheet(name: str):
        ws = wb.create_sheet(title=name[:31])
        _write_header_row(ws, week_cols, header_font, header_fill, header_align, get_column_letter)
        return ws

    # -- MEN ------------------------------------------------------------------
    ws = new_breakdown_sheet("MEN")
    row_ptr = 2
    for leader in [l for l in all_leader_names if _gender_of(l) == "men"]:
        row_ptr, _, _ = _write_leader_block(
            ws, row_ptr, leader, leaders[leader], week_cols,
            label_font, data_align, indent_align,
        )

    # -- WOMEN ------------------------------------------------------------------
    ws = new_breakdown_sheet("WOMEN")
    row_ptr = 2
    for leader in [l for l in all_leader_names if _gender_of(l) == "women"]:
        row_ptr, _, _ = _write_leader_block(
            ws, row_ptr, leader, leaders[leader], week_cols,
            label_font, data_align, indent_align,
        )

    # -- One sheet per named leader — chart ONLY, table hidden -----------------
    for leader_key in MEN_LEADERS + WOMEN_LEADERS:
        matched = next((l for l in all_leader_names if _match_leader(l, leader_key)), None)
        ws = wb.create_sheet(title=leader_key[:31])
        _write_header_row(ws, week_cols, header_font, header_fill, header_align, get_column_letter)

        if not matched:
            # No sessions matched this leader in the given date range.
            # Show a clear message rather than a silently blank tab, and
            # hide the (empty) header table underneath it.
            _hide_table_rows(ws, 1, 1)
            msg = ws.cell(
                row=3, column=1,
                value=f"No cell attendance data found for {leader_key} in this date range.",
            )
            msg.font = Font(italic=True, color="808080")
            continue

        next_row, active_cells_row, totals_row = _write_leader_block(
            ws, 2, matched, leaders[matched], week_cols,
            label_font, data_align, indent_align,
        )
        _add_leader_chart(
            ws, matched, week_cols, anchor_row=next_row + 1,
            active_cells_row=active_cells_row, totals_row=totals_row,
            LineChart=LineChart, Reference=Reference, SeriesLabel=SeriesLabel,
        )
        # Hide the underlying table (header row through the totals row) so
        # only the chart is visible.
        _hide_table_rows(ws, 1, next_row - 1)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()